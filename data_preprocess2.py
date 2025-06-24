import pandas as pd
from openpyxl import load_workbook

source = "jp_eqhist_data_filtered.xlsx"
wb = load_workbook(source)
sheet = wb["Data"]

special_si = {
    "5弱": 5.3,
    "5強": 5.7,
    "6弱": 6.3,
    "6強": 6.7
}

headerrow = sheet[1]

column_index = None
for cell in headerrow:
    if cell.value == "Seismic Intensity":
        column_index = cell.column
        break

# Iterate through the rows and update values in the "Seismic Intensity" column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
    cell = row[0]
    if cell.value in special_si:
        cell.value = special_si[cell.value]

# Save the workbook after updating values
wb.save(source)

# Load the updated Excel file into a pandas DataFrame
df = pd.read_excel(source, sheet_name="Data")

# Convert "Seismic Intensity" column to numeric
df["Seismic Intensity"] = pd.to_numeric(df["Seismic Intensity"], errors="coerce")

# Write the updated DataFrame to the "Data" sheet
for i, row in enumerate(df.itertuples(index=False), start=2):  # Start from row 2 (assuming row 1 is the header)
    for j, value in enumerate(row, start=1):  # Start from column 1
        sheet.cell(row=i, column=j, value=value)

# Save the workbook
wb.save("jp_eqhist_data_filtered.xlsx")