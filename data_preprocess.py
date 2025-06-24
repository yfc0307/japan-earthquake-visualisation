from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

def parse_japanese_datetime(s):
    if not isinstance(s, str):
        return pd.NaT
    s = s.replace("ごろ", "").strip()
    try:
        return datetime.strptime(s, "%Y年%m月%d日 %H時%M分")
    except Exception:
        return pd.NaT

source = "jp_eqhist_data.xlsx"
df = pd.read_excel(source)

df["Time"] = df["Time"].map(parse_japanese_datetime)
df["Magnitude"] = pd.to_numeric(df["Magnitude"], errors="coerce").fillna(0)

# Load the existing workbook
workbook = load_workbook("jp_eqhist_data_filtered.xlsx")

# Select the "Data" sheet
sheet = workbook["Data"]

# Write the updated DataFrame to the "Data" sheet
for i, row in enumerate(df.itertuples(index=False), start=2):  # Start from row 2 (assuming row 1 is the header)
    for j, value in enumerate(row, start=1):  # Start from column 1
        sheet.cell(row=i, column=j, value=value)

# Save the workbook
workbook.save("jp_eqhist_data_filtered.xlsx")